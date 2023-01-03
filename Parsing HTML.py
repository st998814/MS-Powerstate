from asyncio.windows_events import NULL
from itertools import count
from msilib.schema import Class
from pickle import NONE
from tkinter.messagebox import YES
from turtle import width
from urllib.request import urlopen
from bs4 import BeautifulSoup
import re
import json
from cv2 import WARP_INVERSE_MAP
from hamcrest import contains
from regex import F
from sklearn.decomposition import dict_learning
import substring
import pandas as pd
import ast
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
import os

#Launch CMD and gernerate sleepstudy report

os.system('powercfg /sleepstudy')

#Getting raw data

HTMLFile = open(r"C:\Users\st998\Desktop\To Steven\531.html","r")
index = HTMLFile.read()
S = BeautifulSoup(index, 'lxml')
patten = re.compile(r"var LocalSprData = (.*?);$",re.MULTILINE | re.DOTALL)
script = S.find("script", text = patten)
LSD = patten.search(script.prettify()).group(1)
str= LSD
JsonObject= json.loads(LSD)
ScenarioInstances = JsonObject["ScenarioInstances"] #List
sessioncount = ScenarioInstances.count


PS = []

#Structure

class Powerstate():
    def __init__(self,Type,SessionId,EntryTimestampLocal,Duration,OnAc,TopBlockers,swdrip_value,hwdrip_value):
        self.Type = Type
        self.SessionId = SessionId
        self.EntryTimestampLocal = EntryTimestampLocal
        self.Duration = Duration
        self.OnAc = OnAc
        self.TopBlockers = TopBlockers
        self.swdrip_value = swdrip_value
        self.hwdrip_value = hwdrip_value
        



# get data
for instance in ScenarioInstances:
    tYpe = instance["Type"]
    sessionid = instance["SessionId"]
    entrytimestamplocal= instance["EntryTimestampLocal"]
    duration= instance["Duration"]
    onac = instance["OnAc"]
    metadata = instance["Metadata"]["Values"]
    applist=[]
    

   
    if instance.get("TopBlockers")!= None: # 
      
        contains_topblocker= instance["TopBlockers"]
        for x in contains_topblocker:
            if contains_topblocker!=[]:
             applist.append(x["Name"])
   
    swFlag = False
    hwFlag = False
    swdrip_value=0
    hwdrip_value=0
    for k in metadata: #in "metadata"
        key = k["Key"]
        if key== "Info.SwLowPowerStateTime":
            swdrip_value = k["Value"]  # +
            swFlag = True
            
        elif key == "Info.HwLowPowerStateTime":
            hwdrip_value = k["Value"]
            hwFlag = True
            
        if swFlag and hwFlag:   
            break 
        

    Store_to_list= Powerstate(tYpe,sessionid,entrytimestamplocal,duration,onac,applist,swdrip_value,hwdrip_value)
    
    PS.append(Store_to_list)

   



# Find issue

Num = 61100000


Issue_E = []
for i in range(0,len(PS)):
    if PS[i].Duration > 600000000 and PS[i].Type!=2:
        
        Issue_E.append(i+1)
        

Issue_B =[]
for i  in range(0,len(PS)):
    if PS[i].Type==2 and PS[i].swdrip_value==0:
        Issue_B.append(i)
    

        

Issue_C = []
for i  in range(0,len(PS)):
    if PS[i].Type==2 and PS[i].hwdrip_value == 0:
        Issue_C.append(i+1) 
       

Issue_A_swdrip_90= []        #Combine #or
for i  in range(0,len(PS)):

    if PS[i].Type==2 and PS[i].swdrip_value!=0 and PS[i].swdrip_value/Num<90:
    
        Issue_A_swdrip_90.append(i+1)


Issue_A_hwdrip_90 = []
for i  in range(0,len(PS)):
    if PS[i].Type==2 and PS[i].hwdrip_value!=0 and PS[i].hwdrip_value/Num<90:
    
        Issue_A_hwdrip_90.append(i+1)


Issue_F= []

for i in range(0,len(PS)):   #ABS
    if PS[i].swdrip_value/Num - PS[i].hwdrip_value/Num > 10:
     Issue_F.appen(i+1)
for i in range(0,len(PS)):
    if PS[i].hwdrip_value/Num - PS[i].swdrip_value/Num > 10:
           Issue_F.append(i+1)

out_of_wavied_app = []
wavied_app = ["Cortana Voice Activation","Audio Service","PLM Phase Offenders","Maintenance Phase","Host Activity Manager","Windows Error Reporting","DAM Phase Offenders","BI","WNS","NCSI","No CS Phase Offenders","Universal Telemetry Client","DHCP","WP Location Client","BITS Service","WU"]
for i in range(0,len(PS)): 
    if PS[i].TopBlockers == wavied_app:
        out_of_wavied_app.append(i) 



   



# Create excel

df = pd.DataFrame(index=["531"],columns=["Result","Non-Sleep duration over 10 min","No HW Value",'No SW Value',"SW/HW Gap over 0.1"])

if len(Issue_A_swdrip_90) + len(Issue_A_hwdrip_90)+len(Issue_B)+len(Issue_C)+len(Issue_E)+len(Issue_F) == 0:

 df.loc[["531"],['Result']]= "Pass"
else:
 df.loc[["531"],['Result']]= "Fail"

df.loc[["531"],['Non-Sleep duration over 10 min']]= len(Issue_E)
df.loc[["531"],["No SW Value"]] = len(Issue_B)
df.loc[["531"],["No HW Value"]] = len(Issue_C)
df.loc[["531"],["SW/HW Gap over 0.1"]] = len(Issue_F)



df.to_excel("Result_beta.xlsx",index=True,header = True)




wb= openpyxl.load_workbook('Result_beta.xlsx')
ws = wb.active
sheet = wb.worksheets[0]

fill_cell_yellow = PatternFill(patternType='solid', 
                            fgColor='FCBA03') 
fill_cell_green = PatternFill(patternType='solid', 
                            fgColor='35FC03') 
fill_cell_red =PatternFill(patternType='solid', 
                            fgColor='FC2C03')

comment = Comment(text = Issue_E,author = "robot")
comment2 = Comment(text = Issue_B,author = "robot")
comment3 = Comment(text = Issue_C,author = "robot")
comment4 = Comment (text = Issue_F,author = "robot")
ws['C2'].comment = comment
ws['D2'].comment = comment2
ws['E2'].comment = comment3
ws['C1'].fill = fill_cell_yellow
ws['D1'].fill = fill_cell_yellow
ws['E1'].fill = fill_cell_yellow
ws['F1'].fill = fill_cell_yellow

if len(Issue_A_swdrip_90) + len(Issue_A_hwdrip_90)+len(Issue_B)+len(Issue_C)+len(Issue_E)+len(Issue_F) == 0:
    ws['B2'].fill = fill_cell_green
else:
    ws['B2'].fill = fill_cell_red


ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 40


wb.save('Result_beta.xlsx')





out_of_wavied_app = []
wavied_app = ["Cortana Voice Activation","Audio Service","PLM Phase Offenders","Maintenance Phase","Host Activity Manager","Windows Error Reporting","DAM Phase Offenders","BI","WNS","NCSI","No CS Phase Offenders","Universal Telemetry Client","DHCP","WP Location Client","BITS Service","WU"]
for i in range(0,len(PS)): 
    if PS[i].TopBlockers == wavied_app:
        out_of_wavied_app.append(i) 




        





























 










    



 









